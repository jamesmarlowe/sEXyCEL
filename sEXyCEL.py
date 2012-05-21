from openpyxl.cell import get_column_letter
from openpyxl import Workbook

def writeToExcel(iterable,fname):

    wb = Workbook()
    dest_filename = str(fname)+'.xlsx' #name the file
    print dest_filename
    ws = wb.worksheets[0]
    ws.title = "output"                #name the worksheet

    row = 1
    for item in iterable:              #loop through iterable list
        col_idx=1
        try: #list of dictionaries
            for key in item.iterkeys():    #loop through individual iterable items
                col = get_column_letter(col_idx)
                ws.cell('%s%s'%(col, row)).value = str(item[key])    #save data to cell
                col_idx +=1
        except: #list of lists
            for key in item:               #loop through individual iterable items
                col = get_column_letter(col_idx)
                ws.cell('%s%s'%(col, row)).value = str(key)          #save data to cell
                col_idx +=1
        row +=1

    wb.save(filename = dest_filename)
    return
